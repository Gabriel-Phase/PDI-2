import argparse
import datetime
import sys

import cv2
import matplotlib.lines as mlines
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from scipy.optimize import curve_fit


def _gaussian(x, A, x0, w, B):
    return A * np.exp(-2 * ((x - x0) ** 2) / (w ** 2)) + B


def fit_gaussian(x_vals, intensity_vals):
    vals = np.array(intensity_vals, dtype=float)
    x    = np.array(x_vals, dtype=float)

    B0   = float(np.min(vals))
    A0   = float(np.max(vals)) - B0
    x0_0 = float(x[np.argmax(vals)])
    w0   = float(x[-1] - x[0]) / 8

    try:
        popt, _ = curve_fit(_gaussian, x, vals, p0=[A0, x0_0, w0, B0], maxfev=5000)
        A, x0, w, B = popt
        span = x[-1] - x[0]
        if abs(w) < (span / 1000) or abs(w) > span:
            return {'success': False}
        return {'success': True, 'A': A, 'x0': x0, 'w': abs(w), 'B': B}
    except Exception:
        return {'success': False}


def load_grayscale(path):
    img = cv2.imread(path, cv2.IMREAD_UNCHANGED)
    if img is None:
        sys.exit(f"Error: cannot open image '{path}'")
    if img.ndim == 2:
        return img.astype(np.float64)
    if img.shape[2] == 4:
        img = img[:, :, :3]
    b = img[:, :, 0].astype(float)
    g = img[:, :, 1].astype(float)
    r = img[:, :, 2].astype(float)
    return 0.299 * r + 0.587 * g + 0.114 * b


PIXEL_PITCH_UM  = 5.0   # Alvium G1-030 VSWIR sensor pixel pitch (µm)
OBJECTIVE_MAG   = 100   # Mitutoyo 100x M Plan APO NIR
ZOOM_MAG        = 2.75  # Optem Fusion 7:1 at maximum zoom (range: 0.75x–5.25x)
CAMERA_TUBE_MAG = 1.0   # Optem Fusion camera tube 35-08-06-000

PIXEL_SIZE_UM = PIXEL_PITCH_UM / (OBJECTIVE_MAG * ZOOM_MAG * CAMERA_TUBE_MAG)


def export_excel(path, date_str, pixel_size_mm,
                 x_mm, row, fit_x,
                 y_mm, col, fit_y):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    DARK_BLUE  = '1F3864'
    MID_BLUE   = '2E75B6'
    LIGHT_BLUE = 'D6E4F0'
    WHITE      = 'FFFFFF'
    LIGHT_GRAY = 'F2F2F2'

    def header_font(bold=True, color=WHITE, size=11):
        return Font(name='Calibri', bold=bold, color=color, size=size)

    def cell_font(bold=False, color='000000', size=10):
        return Font(name='Calibri', bold=bold, color=color, size=size)

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def thin_border():
        s = Side(style='thin', color='BFBFBF')
        return Border(left=s, right=s, top=s, bottom=s)

    def style_header_row(ws, row_num, col_count, bg=MID_BLUE):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font      = header_font()
            cell.fill      = fill(bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = thin_border()

    def style_data_row(ws, row_num, col_count, alternate=False):
        bg = LIGHT_GRAY if alternate else WHITE
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font      = cell_font()
            cell.fill      = fill(bg)
            cell.alignment = Alignment(horizontal='left' if c == 1 else 'center',
                                       vertical='center')
            cell.border    = thin_border()

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = openpyxl.Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    # Title banner
    ws.merge_cells('A1:C1')
    title_cell = ws['A1']
    title_cell.value     = 'POP Beam Profile  —  Measurement Summary'
    title_cell.font      = Font(name='Calibri', bold=True, color=WHITE, size=13)
    title_cell.fill      = fill(DARK_BLUE)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.append([])  # spacer row

    # Meta info
    for label, value in [('Date', date_str), ('Pixel size (µm/px)', pixel_size_mm * 1e3)]:
        ws.append([label, value])
        r = ws.max_row
        ws.cell(r, 1).font      = cell_font(bold=True)
        ws.cell(r, 1).fill      = fill(LIGHT_BLUE)
        ws.cell(r, 1).border    = thin_border()
        ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(r, 2).font      = cell_font()
        ws.cell(r, 2).fill      = fill(WHITE)
        ws.cell(r, 2).border    = thin_border()
        ws.cell(r, 2).alignment = Alignment(horizontal='left', vertical='center')

    ws.append([])  # spacer

    # Column headers
    ws.append(['Parameter', 'X Scan', 'Y Scan'])
    style_header_row(ws, ws.max_row, 3)

    # Data rows
    diam_x = 2 * fit_x['w'] * 1e3 if fit_x['success'] else 'N/A'
    diam_y = 2 * fit_y['w'] * 1e3 if fit_y['success'] else 'N/A'
    data_rows = [
        ('Fit success',        fit_x['success'],          fit_y['success']),
        ('Amplitude (A)',      fit_x.get('A',  'N/A'),    fit_y.get('A',  'N/A')),
        ('Center x0 (mm)',     fit_x.get('x0', 'N/A'),    fit_y.get('x0', 'N/A')),
        ('1/e² Radius w (µm)', fit_x.get('w', 'N/A') * 1e3 if fit_x.get('w') is not None else 'N/A',
                               fit_y.get('w', 'N/A') * 1e3 if fit_y.get('w') is not None else 'N/A'),
        ('Baseline (B)',       fit_x.get('B',  'N/A'),    fit_y.get('B',  'N/A')),
        ('1/e² Diameter (µm)', diam_x,                    diam_y),
    ]
    for i, (label, xv, yv) in enumerate(data_rows):
        ws.append([label, xv, yv])
        r = ws.max_row
        style_data_row(ws, r, 3, alternate=(i % 2 == 0))
        if label == '1/e² Diameter (µm)':
            for c in range(1, 4):
                ws.cell(r, c).font = cell_font(bold=True, color=DARK_BLUE)
                ws.cell(r, c).fill = fill(LIGHT_BLUE)
        for c in (2, 3):
            cell = ws.cell(r, c)
            if isinstance(cell.value, float):
                cell.number_format = '0.000000'

    set_col_widths(ws, [28, 18, 18])

    # ── Scan data sheets ─────────────────────────────────────────────────────
    for sheet_title, positions, intensities in [
        ('X Scan', x_mm, row),
        ('Y Scan', y_mm, col),
    ]:
        color = '1F497D' if sheet_title == 'X Scan' else '375623'
        wsh = wb.create_sheet(sheet_title)
        wsh.sheet_view.showGridLines = False

        wsh.append(['Position (mm)', 'Intensity'])
        style_header_row(wsh, 1, 2, bg=color)
        wsh.row_dimensions[1].height = 20

        for i, (pv, iv) in enumerate(zip(positions.tolist(), intensities.tolist())):
            wsh.append([pv, iv])
            r = wsh.max_row
            bg = LIGHT_GRAY if i % 2 else WHITE
            for c in (1, 2):
                cell = wsh.cell(r, c)
                cell.font         = cell_font(size=9)
                cell.fill         = fill(bg)
                cell.border       = thin_border()
                cell.alignment    = Alignment(horizontal='center', vertical='center')
                cell.number_format = '0.000000'

        set_col_widths(wsh, [18, 18])

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description='Beam profile line scan from image')
    parser.add_argument('image', help='Path to input image')
    parser.add_argument('--pixel-size', type=float, default=None,
                        help='Override pixel size in µm/pixel (default: calculated from optics)')
    args = parser.parse_args()

    gray = load_grayscale(args.image)
    h, w = gray.shape

    row = gray[h // 2]
    col = gray[:, w // 2]

    pixel_size_mm = (args.pixel_size if args.pixel_size is not None else PIXEL_SIZE_UM) * 1e-3
    x_mm = (np.arange(w) - w / 2) * pixel_size_mm
    y_mm = (np.arange(h) - h / 2) * pixel_size_mm

    fit_x = fit_gaussian(x_mm, row)
    fit_y = fit_gaussian(y_mm, col)

    today = datetime.date.today()
    date_str = f"{today.month}/{today.day}/{today.year}"

    if fit_x['success']:
        diam_x_str = f"1/e^2 Gaussian Diameter (X) = {2 * fit_x['w'] * 1e3:.4f} um"
    else:
        diam_x_str = "Fit X failed"

    if fit_y['success']:
        diam_y_str = f"1/e^2 Gaussian Diameter (Y) = {2 * fit_y['w'] * 1e3:.4f} um"
    else:
        diam_y_str = "Fit Y failed"

    fig = plt.figure(figsize=(14, 5))
    plot_bottom = 0.42

    ax_x = fig.add_axes([0.07, plot_bottom, 0.38, 0.52])
    ax_y = fig.add_axes([0.55, plot_bottom, 0.38, 0.52])

    ax_x.plot(x_mm, row, color='blue', linewidth=1)
    if fit_x['success']:
        level_x = fit_x['A'] / np.e ** 2 + fit_x['B']
        ax_x.axhline(level_x, color='red', linestyle='--', linewidth=1)
    ax_x.set_title('X Scan', fontsize=11)
    ax_x.set_xlabel('Position (mm)', fontsize=10)
    ax_x.set_ylabel('Intensity (a.u.)', fontsize=10)
    ax_x.grid(True, alpha=0.4)
    ax_x.set_xlim(x_mm[0], x_mm[-1])

    ax_y.plot(y_mm, col, color='green', linewidth=1)
    if fit_y['success']:
        level_y = fit_y['A'] / np.e ** 2 + fit_y['B']
        ax_y.axhline(level_y, color='red', linestyle='--', linewidth=1)
    ax_y.set_title('Y Scan', fontsize=11)
    ax_y.set_xlabel('Position (mm)', fontsize=10)
    ax_y.set_ylabel('Intensity (a.u.)', fontsize=10)
    ax_y.grid(True, alpha=0.4)
    ax_y.set_xlim(y_mm[0], y_mm[-1])

    fig.text(0.5, 0.215, 'POP beam profile', ha='center', va='top',
             fontsize=10, family='monospace')

    fig.add_artist(mlines.Line2D([0.05, 0.95], [0.175, 0.175],
                                  transform=fig.transFigure, color='black', linewidth=0.8))

    fig.text(0.05, 0.145, date_str,    ha='left', va='top', fontsize=9, family='monospace')
    fig.text(0.05, 0.10,  diam_x_str, ha='left', va='top', fontsize=9, family='monospace')
    fig.text(0.05, 0.06,  diam_y_str, ha='left', va='top', fontsize=9, family='monospace')

    export_excel('beam_profile.xlsx', date_str, pixel_size_mm,
                 x_mm, row, fit_x,
                 y_mm, col, fit_y)

    plt.savefig('beam_profile.png', dpi=150, bbox_inches='tight')
    plt.show()



if __name__ == '__main__':
    main()
# python image_scan.py photos\image.png